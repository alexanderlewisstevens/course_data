#!/usr/bin/env python3
"""
Build per-term workbooks:
- Master_PQ.xlsx: GTA Eligible + Time Conflicts (read-only source)
- Faculty_Preferences.xlsx: full master with User Edits (generated via build_term_excel)
- GTA_Editable.xlsx: GTA feed columns (no Prefer TA), editable fields unlocked (In Class, Office Hours, Grading, Time commitment, Notes).
"""

from __future__ import annotations

import argparse
import pathlib
import sys

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.build_term_excel import build_workbook, load_conflict_csv, load_gta_json, TEMPLATE_HEADERS  # type: ignore
from scripts.build_gta_feed import FIELDS as GTA_FIELDS, load_gta_json as load_gta_feed_json  # type: ignore


def build_master(term: str, processed_dir: pathlib.Path, export_dir: pathlib.Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GTA Eligible"
    ws.append(TEMPLATE_HEADERS)
    for r in load_gta_json(term, processed_dir):
        ws.append(
            [
                term,
                r.get("crn", ""),
                r.get("course", ""),
                r.get("section", ""),
                r.get("title", ""),
                r.get("course_type", ""),
                r.get("meeting_dates", ""),
                r.get("time", ""),
                r.get("days", ""),
                r.get("hours", ""),
                r.get("room", ""),
                r.get("instructor", ""),
                r.get("seats", ""),
                r.get("enrolled", ""),
                r.get("crosslisted_enrollment", ""),
                r.get("total_enrollment", ""),
                "", "", "", "", "", "", "", "", "",
            ]
        )
    ws.protection.sheet = True
    # Time Conflicts
    ws2 = wb.create_sheet("Time Conflicts")
    for row in load_conflict_csv(term, processed_dir):
        ws2.append(row)
    out = export_dir / "Master_PQ.xlsx"
    export_dir.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote master workbook: {out}")


def build_gta_editable(term: str, processed_dir: pathlib.Path, export_dir: pathlib.Path) -> None:
    rows = load_gta_feed_json(term, processed_dir)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GTA Editable"
    ws.append(GTA_FIELDS)
    for r in rows:
        ws.append(
            [
                term,
                r.get("crn", ""),
                r.get("course", ""),
                r.get("section", ""),
                r.get("title", ""),
                r.get("course_type", ""),
                r.get("meeting_dates", ""),
                r.get("time", ""),
                r.get("days", ""),
                r.get("hours", ""),
                r.get("room", ""),
                r.get("instructor", ""),
                r.get("seats", ""),
                r.get("enrolled", ""),
                r.get("crosslisted_enrollment", ""),
                r.get("total_enrollment", ""),
                "", "", "", "", "",
            ]
        )
    # Unlock editable columns (In Class onward)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=17, max_col=ws.max_column):
        for cell in row:
            cell.protection = cell.protection.copy(locked=False)
    ws.protection.sheet = True
    out = export_dir / "GTA_Editable.xlsx"
    export_dir.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote GTA editable workbook: {out}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build per-term workbooks (master, faculty, GTA).")
    parser.add_argument("terms", nargs="+", help="Term codes, e.g., 202610")
    parser.add_argument("--processed-dir", default="data/processed", help="Processed data directory")
    parser.add_argument("--export-dir", default="data/exports", help="Exports root directory")
    args = parser.parse_args()

    processed_dir = pathlib.Path(args.processed_dir)
    export_root = pathlib.Path(args.export_dir)

    for term in args.terms:
        term_dir = export_root / term
        term_dir.mkdir(parents=True, exist_ok=True)
        build_master(term, processed_dir, term_dir)
        # Faculty workbook via existing builder
        build_workbook(term, load_gta_json(term, processed_dir), load_conflict_csv(term, processed_dir), term_dir / "Faculty_Preferences.xlsx")
        build_gta_editable(term, processed_dir, term_dir)


if __name__ == "__main__":
    main()
