#!/usr/bin/env python3
"""
Normalize legacy instructor preference Excel files into the current template shape.

Reads all .xlsx/.xlsm files under an input directory (default: data/history/excel/)
and writes normalized copies (default: data/history/normalized/) with a single sheet
named "Faculty Preferences" and headers matching the Faculty GTA Survey template.

Target columns (in order):
CRN, Course, Sec, Title, Course Type, Meeting Dates, Time, Days, Hrs, Room,
Instructor, Seat, Enr, Crosslisted Enr, Total Enr, Prefer TA 1, Prefer TA 2,
Prefer TA 3, Prefer TA 4, In Class, Office Hours, Grading, Time commitment, Notes
"""

from __future__ import annotations

import argparse
import pathlib
import re
from typing import Dict, List, Any

import openpyxl

TARGET_HEADERS = [
    "Term",
    "CRN",
    "Course",
    "Sec",
    "Title",
    "Course Type",
    "Meeting Dates",
    "Time",
    "Days",
    "Hrs",
    "Room",
    "Instructor",
    "Seat",
    "Enr",
    "Crosslisted Enr",
    "Total Enr",
    "Prefer TA 1",
    "Prefer TA 2",
    "Prefer TA 3",
    "In Class",
    "Office Hours",
    "Grading",
    "Time commitment",
    "Notes",
]

HEADER_MAP: Dict[str, str] = {
    "term": "Term",
    "term code": "Term",
    "term_code": "Term",
    "crn": "CRN",
    "course": "Course",
    "course number": "Course",
    "course_number": "Course",
    "course num": "Course",
    "course_num": "Course",
    "sec": "Sec",
    "section": "Sec",
    "title": "Title",
    "course type": "Course Type",
    "course_type": "Course Type",
    "meeting dates": "Meeting Dates",
    "meeting_dates": "Meeting Dates",
    "time": "Time",
    "days": "Days",
    "hrs": "Hrs",
    "hours": "Hrs",
    "room": "Room",
    "instructor": "Instructor",
    "listed instructor": "Instructor",
    "listed_instructor": "Instructor",
    "updated instructor": "Instructor",
    "updated_instructor": "Instructor",
    "seat": "Seat",
    "seats": "Seat",
    "enr": "Enr",
    "enrolled": "Enr",
    "crosslisted enr": "Crosslisted Enr",
    "cross listed": "Crosslisted Enr",
    "crosslisted": "Crosslisted Enr",
    "total enr": "Total Enr",
    "total enrollment": "Total Enr",
    "prefer ta 1": "Prefer TA 1",
    "prefer ta 2": "Prefer TA 2",
    "prefer ta 3": "Prefer TA 3",
    "in class": "In Class",
    "in-class": "In Class",
    "office hours": "Office Hours",
    "office_hours": "Office Hours",
    "grading": "Grading",
    "time commitment": "Time commitment",
    "time committment": "Time commitment",
    "time_commitment": "Time commitment",
    "notes": "Notes",
    "course notes": "Notes",
    "course_notes": "Notes",
}


def _normalize_instructor(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return ""
    if "," in name:
        return name
    parts = name.split()
    if len(parts) == 1:
        return name
    last = parts[-1]
    first = " ".join(parts[:-1])
    return f"{last}, {first}"


def _normalize_course(course: str) -> str:
    course = (course or "").strip()
    if not course:
        return ""
    m = re.match(r"^([A-Za-z]+)\s*(\d+)$", course)
    if m:
        return f"{m.group(1).upper()} {m.group(2)}"
    if re.match(r"^\d+$", course):
        return f"COMP {course}"
    return course


def _infer_term_code(path: pathlib.Path) -> str:
    name = path.stem
    m = re.search(r"(winter|spring|summer|autumn|fall)\s*(\d{4})", name, re.IGNORECASE)
    if not m:
        return ""
    season = m.group(1).lower()
    year = m.group(2)
    offset = {"winter": "10", "spring": "30", "summer": "50", "autumn": "70", "fall": "70"}.get(season, "")
    if not offset:
        return ""
    return f"{year}{offset}"


def normalize_file(src: pathlib.Path, dest_dir: pathlib.Path) -> pathlib.Path | None:
    try:
        wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    except Exception as exc:
        print(f"Skipping {src.name}: cannot open ({exc})")
        return None

    out_rows: List[List[Any]] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        # Build a map from target header -> column index
        colmap: Dict[str, int] = {}
        for idx, h in enumerate(header):
            if h is None:
                continue
            key = str(h).strip().lower()
            target = HEADER_MAP.get(key)
            if target:
                # Prefer updated over listed if both present; later overwrite earlier
                colmap[target] = idx
        # Must at least have Course and Instructor to consider this sheet
        if "Course" not in colmap or "Instructor" not in colmap:
            continue
        for row in rows[1:]:
            if row is None:
                continue
            term_val = ""
            if "Term" in colmap:
                term_idx = colmap["Term"]
                term_val = str(row[term_idx]).strip() if term_idx is not None and row[term_idx] is not None else ""
            if not term_val:
                term_val = _infer_term_code(src)
            course_val = str(row[colmap["Course"]]).strip() if colmap.get("Course") is not None and row[colmap["Course"]] is not None else ""
            course_val = _normalize_course(course_val)
            instr_val_raw = row[colmap["Instructor"]] if colmap.get("Instructor") is not None and row[colmap["Instructor"]] is not None else ""
            instr_val = _normalize_instructor(str(instr_val_raw))
            if not course_val and not instr_val:
                continue
            out_row: List[Any] = []
            for target in TARGET_HEADERS:
                idx = colmap.get(target)
                val = row[idx] if idx is not None and idx < len(row) else ""
                if target == "Term":
                    val = term_val
                if target == "Course":
                    val = course_val
                if target == "Instructor":
                    val = instr_val
                if target == "Title" and val:
                    val = str(val).strip()
                if target.startswith("Prefer TA"):
                    val = ""
                out_row.append(val if val is not None else "")
            out_rows.append(out_row)

    if not out_rows:
        print(f"Skipping {src.name}: no usable rows found")
        return None

    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / src.name
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Faculty Preferences"
    out_ws.append(TARGET_HEADERS)
    for r in out_rows:
        out_ws.append(r)
    out_wb.save(dest_path)
    return dest_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize legacy instructor preference Excel files to the standard template.")
    parser.add_argument("--input-dir", default="data/history/excel", help="Directory containing legacy Excel files.")
    parser.add_argument("--output-dir", default="data/history/normalized", help="Directory to write normalized Excel files.")
    args = parser.parse_args()

    input_dir = pathlib.Path(args.input_dir)
    output_dir = pathlib.Path(args.output_dir)
    files = [p for p in input_dir.glob("**/*") if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}]
    if not files:
        print(f"No Excel files found in {input_dir}")
        return

    written = 0
    for f in sorted(files):
        dest = normalize_file(f, output_dir)
        if dest:
            written += 1
            print(f"Normalized {f.name} -> {dest}")
    print(f"Completed. Wrote {written} normalized file(s) to {output_dir}.")


if __name__ == "__main__":
    main()
