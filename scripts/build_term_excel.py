#!/usr/bin/env python3
"""
Build a term-specific Excel workbook with:
- Sheet "GTA Eligible" populated from gta_eligible_term_<term>.json using the template headers.
- Sheet "Time Conflicts" populated from gta_compatibility_term_<term>.csv.

Usage:
    python scripts/build_term_excel.py 202610 --output data/exports/term_202610_gta.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import pathlib
import re
import sys
from typing import List

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

TEMPLATE_HEADERS = [
    "Term",
    "CRN",
    "Course",
    "Sec",
    "Title",
    "Course Type",
    "Meeting Dates",
    "Time",
    "Days",
    "Hrs",
    "Room",
    "Instructor",
    "Seat",
    "Enr",
    "Crosslisted Enr",
    "Total Enr",
    "Prefer TA 1",
    "Prefer TA 2",
    "Prefer TA 3",
    "In Class",
    "Office Hours",
    "Grading",
    "Time commitment",
    "Notes",
]

TEMPLATE_PATH = pathlib.Path("data/history/templates/Faculty GTA Survey template.xlsx")


def load_template_name_headers(template_path: pathlib.Path = TEMPLATE_PATH) -> List[str]:
    if not template_path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    except Exception:
        return []
    ws = wb["GTA info"] if "GTA info" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 24:
        return []
    raw = rows[23]
    cleaned = []
    for v in raw:
        if v is None:
            continue
        s = str(v).strip().strip('"').replace("\xa0", " ")
        if s:
            cleaned.append(s)
    return cleaned


def load_gta_json(term: str, base_dir: pathlib.Path) -> List[dict]:
    path = base_dir / f"gta_eligible_term_{term}.json"
    data = json.loads(path.read_text())
    return data


def load_conflict_csv(term: str, base_dir: pathlib.Path) -> List[List[str]]:
    path = base_dir / f"gta_compatibility_term_{term}.csv"
    with path.open() as f:
        reader = csv.reader(f)
        return [row for row in reader]


def build_workbook(term: str, gta_rows: List[dict], conflicts: List[List[str]], output_path: pathlib.Path) -> None:
    wb = openpyxl.Workbook()

    # GTA Eligible sheet
    ws = wb.active
    ws.title = "GTA Eligible"
    ws.append(TEMPLATE_HEADERS)
    for r in gta_rows:
        row = [
            term,
            r.get("crn", ""),
            r.get("course", ""),
            r.get("section", ""),
            r.get("title", ""),
            r.get("course_type", ""),
            r.get("meeting_dates", ""),
            r.get("time", ""),
            r.get("days", ""),
            r.get("hours", ""),
            r.get("room", ""),
            r.get("instructor", ""),
            r.get("seats", ""),
            r.get("enrolled", ""),
            r.get("crosslisted_enrollment", ""),
            r.get("total_enrollment", ""),
            "",  # Prefer TA 1
            "",  # Prefer TA 2
            "",  # Prefer TA 3
            "",  # In Class
            "",  # Office Hours
            "",  # Grading
            "",  # Time commitment
            "",  # Notes
        ]
        ws.append(row)

    # GTA Names Headers sheet
    names = load_template_name_headers()
    names_ws = wb.create_sheet("Faculty Preferences")
    if names:
        names_ws.append(names)

    # User Edits sheet: locked columns A-P, editable beyond; formulas pull from GTA Eligible by CRN
    edits_ws = wb.create_sheet("User Edits")
    edits_ws.append(TEMPLATE_HEADERS)
    n_rows = len(gta_rows)
    header_to_col_letter = {h: openpyxl.utils.get_column_letter(idx + 1) for idx, h in enumerate(TEMPLATE_HEADERS)}
    crn_range = f"'GTA Eligible'!$B$2:$B${n_rows+1}"
    for i in range(2, n_rows + 2):
        # CRN direct copy
        crn_val = ws.cell(row=i, column=2).value
        edits_ws.cell(row=i, column=2, value=crn_val)
        for idx, header in enumerate(TEMPLATE_HEADERS, start=1):
            cell = edits_ws.cell(row=i, column=idx)
            if idx == 2:
                continue
            if idx <= 16:
                col_letter = header_to_col_letter[header]
                cell.value = f"=XLOOKUP($B{ i },{crn_range},'GTA Eligible'!${col_letter}$2:${col_letter}${n_rows+1},\"\")"
            else:
                cell.protection = cell.protection.copy(locked=False)
        # Unlock editable columns explicitly
        for idx in range(17, len(TEMPLATE_HEADERS) + 1):
            cell = edits_ws.cell(row=i, column=idx)
            cell.protection = cell.protection.copy(locked=False)
    edits_ws.protection.sheet = True

    # Time Conflicts sheet
    ws2 = wb.create_sheet("Time Conflicts")
    for row in conflicts:
        ws2.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote workbook: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build term Excel with GTA eligible and conflict matrix.")
    parser.add_argument("term", help="Term code, e.g., 202610")
    parser.add_argument("--processed-dir", default="data/processed", help="Directory containing processed JSON/CSV.")
    parser.add_argument("--output", default="", help="Output Excel path. Default: data/exports/term_<term>_gta.xlsx")
    args = parser.parse_args()

    processed_dir = pathlib.Path(args.processed_dir)
    if args.output:
        output_path = pathlib.Path(args.output)
    else:
        try:
            import config  # type: ignore
            label = getattr(config, "TERMS", {}).get(args.term, args.term)
        except Exception:
            label = args.term
        parts = label.split()
        season = parts[0] if parts else ""
        year = parts[-1] if parts and parts[-1].isdigit() else ""
        if season and year:
            label_safe = re.sub(r"[^A-Za-z0-9]+", "_", f"{year}_{season}").strip("_")
        else:
            label_safe = re.sub(r"[^A-Za-z0-9]+", "_", label).strip("_")
        name_parts = [args.term, "Faculty_Preferences", label_safe]
        output_path = pathlib.Path(f"data/exports/{'_'.join(p for p in name_parts if p)}.xlsx")

    gta_rows = load_gta_json(args.term, processed_dir)
    conflicts = load_conflict_csv(args.term, processed_dir)
    build_workbook(args.term, gta_rows, conflicts, output_path)


if __name__ == "__main__":
    main()
